from flaskr import app
from flask import render_template, request, redirect, url_for, send_file
from io import BytesIO
from openpyxl.styles.fonts import Font
import requests
import configparser
import datetime
import openpyxl as excel

# INIファイル名を設定する
INI_FILE_NAME = 'config.ini'

@app.route('/')
def index():
  items = []
  d_param = {}
  return render_template(
    'index.html',
    items=items,
    d_param=d_param
  )
  
@app.route('/search', methods={'POST'})
def search():
  d_param = {}
  #ID 設定
  d_param['app_id'] = get_value('DEFAULT', 'apikey')

  #検索キーワード
  if request.form['keyword'] != "":
    d_param['keyword'] = request.form['keyword']
  else:
    #設定なければ、ゼロ件設定し画面に戻る
    items = []
    return render_template(
      'index.html',
      items=items,
      d_param=d_param
    )
    
  # その他入力値
  d_param['sort'] = request.form['sort']
  d_param['minPrice'] = request.form['minPrice']
  d_param['maxPrice'] = request.form['maxPrice']

  # 楽天データ取得
  r_json = get_rakudata(d_param)
  
  items = []
  item = {}
  if r_json["count"] > 0:
    for i in r_json["Items"]:

      item = i["Item"]
      
      item = {'code' : item["itemCode"],
        'name' : item["itemName"],
        'price' : item["itemPrice"],
        'point' : item["pointRate"],
        'url' : item["itemUrl"],
        'shop' : item['shopName']}
      items.append(item)

  return render_template(
    'index.html',
    items=items,
    d_param=d_param
  )

# ダウンロード機能
@app.route('/download', methods={'POST'})
def download():
  d_param = {}
  #ID 設定
  d_param['app_id'] = get_value('DEFAULT', 'apikey')

  #検索キーワード
  if request.form['keyword'] != "":
    d_param['keyword'] = request.form['keyword']
  else:
    #設定なければ、ゼロ件設定し画面に戻る
    items = []
    return render_template(
      'index.html',
      items=items,
      d_param=d_param
    )
    
  # その他入力値
  d_param['sort'] = request.form['sort']
  d_param['minPrice'] = request.form['minPrice']
  d_param['maxPrice'] = request.form['maxPrice']

  # 楽天データ取得
  r_json = get_rakudata(d_param)

  # 出力先ストリームの設定
  output = BytesIO()

  # テンプレートファイルを読み込み
  wb = excel.load_workbook('excel_template.xlsx')
  ws = wb.active

  # ヘッダーの書き込み 指定のセルに部署名と名前を書き込む
  ws.cell(1, 1, value='商品コード')
  ws.cell(1, 2, value='タイトル')
  ws.cell(1, 3, value='金額')
  ws.cell(1, 4, value='ポイント倍付け')
  ws.cell(1, 5, value='店名')
  
  if r_json["count"] > 0:
    j = 1
    for i in r_json["Items"]:
      j += 1
      item = i["Item"]
      ws.cell(j, 1).value = item["itemCode"]
      ws.cell(j, 2).value = item["itemName"]
      ws.cell(j, 2).hyperlink = item["itemUrl"]
      ws.cell(j, 2).font = Font(color='0000FF', u='single')
      ws.cell(j, 3).value = item["itemPrice"]
      ws.cell(j, 4).value = item["pointRate"]
      ws.cell(j, 5).value = item['shopName']
  
  #ワークブックをメモリ上へ出力
  wb.save(output)
  output.seek(0)
  wb.close()

  # 書き込みファイル名は時間を使用
  dt_now = datetime.datetime.now()
  name = 'rakudata_{date}.xlsx'.format(date=dt_now.strftime('%Y%m%d%H%M%S'))
    
  # ブラウザへExcelファイルを設定
  return send_file(output, download_name=name, as_attachment=True)

# INIファイル取得  
def get_value(section, key):
    # ConfigParserオブジェクトを生成する
    config = configparser.ConfigParser()
    
    # INIファイルを読み込む
    config.read(INI_FILE_NAME, encoding='utf-8')
    
    # 指定されたセクションとキーに対応する値を返す
    return config[section][key]

# 楽天データ取得
def get_rakudata(d_param):
  # URL
  url = "https://app.rakuten.co.jp/services/api/IchibaItem/Search/20220601"

  # パラメータ設定
  payload = {
    "keyword":d_param['keyword'],
    "applicationId":d_param['app_id'],
    "hits":20,
    "page":1,
    "minPrice":d_param['minPrice'],
    "maxPrice":d_param['maxPrice'],
    "sort":d_param['sort'],
    "postageFlag":1 #送料込み
  }

  r = requests.get(url, params=payload)
  r_json = r.json()
  
  return r_json